import pandas as pd
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import json


class DataProcessor:
    def __init__(self, column_mapping=None):
        self.column_mapping = column_mapping or {}

    def apply_column_mapping(self, df):
        """应用自定义列头映射"""
        if df is None or df.empty or not self.column_mapping:
            return df
        
        # 创建新的列名映射
        new_columns = {}
        for old_col in df.columns:
            new_col = self.column_mapping.get(old_col, old_col)
            new_columns[old_col] = new_col
        
        # 重命名列
        df = df.rename(columns=new_columns)
        return df
    
    def clean_data(self, df):
        """清洗数据"""
        if df is None or df.empty:
            return pd.DataFrame()
        
        # 应用列头映射
        df = self.apply_column_mapping(df)
        
        # 移除空行
        df = df.dropna(how='all')
        
        # 清理字符串数据
        for col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].astype(str).str.strip()
                # 将空字符串替换为NaN
                df[col] = df[col].replace('', '-')
                df[col] = df[col].replace('nan', '-')
        
        return df

    def save_to_excel(self, df, filename=None):
        """保存数据到Excel文件，保留超链接"""
        if df is None or df.empty:
            print("警告：数据为空，无法保存")
            return None
            
        if not filename:
            filename = f"givemeoc_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        elif not filename.endswith('.xlsx'):
            filename = f"{filename}.xlsx"
        
        # 确保输出目录存在
        os.makedirs('data', exist_ok=True)
        filepath = os.path.join('data', filename)
        
        try:
            # 创建Excel工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "招聘信息"
            
            # 写入表头
            for col_idx, header in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 写入数据并设置超链接
            for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    
                    # 检查是否为URL
                    if isinstance(value, str) and value.startswith('http'):
                        cell.value = value
                        #cell.hyperlink = value
                        cell.font = Font(color="0000FF", underline="single")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.value = value if str(value) != 'nan' and str(value) != '-' else ''
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # 调整列宽
            for col_idx, column in enumerate(df.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                
                # 检查表头长度
                header_length = len(str(column))
                max_length = max(max_length, header_length)
                
                # 检查数据长度
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            max_length = max(max_length, cell_length)
                
                # 设置列宽，限制最大宽度
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 设置行高
            for row in ws.iter_rows():
                ws.row_dimensions[row[0].row].height = 20
            
            wb.save(filepath)
            print(f"数据已保存到Excel文件: {filepath}")
            
            # 显示超链接统计
            url_columns = [col for col in df.columns if df[col].astype(str).str.startswith('http').any()]
            if url_columns:
                print(f"以下列包含超链接: {', '.join(url_columns)}")
            
            return filepath
            
        except Exception as e:
            print(f"保存Excel文件时出错: {str(e)}")
            return None

    def save_to_json(self, df, filename=None):
        """保存数据到JSON文件"""
        if df is None or df.empty:
            print("警告：数据为空，无法保存")
            return None
            
        if not filename:
            filename = f"givemeoc_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        elif not filename.endswith('.json'):
            filename = f"{filename}.json"
        
        # 确保输出目录存在
        os.makedirs('data', exist_ok=True)
        filepath = os.path.join('data', filename)
        
        try:
            # 转换为字典格式
            data = df.to_dict('records')
            
            # 保存为JSON
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            print(f"数据已保存到JSON文件: {filepath}")
            return filepath
            
        except Exception as e:
            print(f"保存JSON文件时出错: {str(e)}")
            return None